import openpyxl
from openpyxl.cell import get_column_letter, column_index_from_string
from optparse import OptionParser


wb = openpyxl.Workbook()
sheet = wb.get_active_sheet()


class Multiplication_table():
    def __init__(self, sheet):
        self.sheet = sheet

    def make_header(self, num, start=1):
        for i in range(start, num+1):
            self.sheet[get_column_letter(i+1) + '1'] = i
            self.sheet['A' + str(i+1)] = i
    
    def fill_content(self, num, start=1):
        for i in range(start, num+1):
            for j in range(start, num+1):
                self.sheet[get_column_letter(i+1) + str(j+1)] = i*j

                
def main():
    parser = OptionParser(usage='usage: %prog -n <num_of_matrix>', 
                          version='%prog 1.0')
    parser.add_option('-n', 
                      dest='NUM_OF_MATRIX', 
                      type='int', 
                      help="number of matrix")
    
    (options, args) = parser.parse_args()
    if not options.NUM_OF_MATRIX:
        parser.error('Wrong arguments')
    
    mt = Multiplication_table(sheet)
    mt.make_header(options.NUM_OF_MATRIX)
    mt.fill_content(options.NUM_OF_MATRIX)
    wb.save('multiplicationtable.xlsx')
    
    
if __name__ == '__main__':
    main()