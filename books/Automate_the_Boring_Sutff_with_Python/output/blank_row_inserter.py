import openpyxl
from optparse import OptionParser


def main():
    parser = OptionParser(usage='%prog -n <start_row> -m <num_of_blank>', 
                          version='%prog 1.0')
    parser.add_option('-N', 
                      dest='start_row', 
                      type='int', 
                      help="start row for to insert row")
    
    parser.add_option('-M', 
                  dest='num_of_blank', 
                  type='int', 
                  help="To insert blank rows at start row")
    
    parser.add_option('-n',
                      dest='name',
                      type='string',
                      help='Excel name')

    
    (options, args) = parser.parse_args()
    if not options.start_row or not options.num_of_blank or not options.name:
        parser.error('Wrong arguments')
    
    wb = openpyxl.load_workbook(options.name)
    sheet = wb.get_active_sheet()
    
    wb2 = openpyxl.Workbook()
    sheet2 = wb2.get_active_sheet()
    
    for rowNum in range(1, options.start_row+1):
        for colNum in range(1, sheet.get_highest_column()+1):
            sheet2.cell(row=rowNum, column=colNum).value = sheet.cell(row=rowNum, column=colNum).value
    
    
    for rowNum in range(options.start_row, sheet.get_highest_row()+1):
        for colNum in range(1, sheet.get_highest_column()+1):
            sheet2.cell(row=rowNum+options.num_of_blank, column=colNum).value = \
            sheet.cell(row=rowNum, column=colNum).value
    
    wb2.save('added_blank.xlsx')
    
    
if __name__ == '__main__':
    main()